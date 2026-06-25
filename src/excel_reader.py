from datetime import datetime, date, time
import openpyxl

class UnrecognizedFormatError(Exception):
    pass

META = {
    "מספר עובד": "employee_no",
    "שם פרטי": "first_name",
    "שם משפחה": "last_name",
    "עיר": "city",
    "הסעה": "route",
}
LATE_HDR = "זמן איחור"
ARRIVAL_HDR = "זמן הגעה"
_PY2SHORT = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]

def _weekday(d):
    return _PY2SHORT[d.weekday()]

def _fmt_time(v):
    if isinstance(v, (datetime, time)):
        return v.strftime("%H:%M")
    return None

def _as_date(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return None

def _find_sheet(wb):
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))
        if rows and any(c in META for c in rows[0] if c is not None):
            return ws
    raise UnrecognizedFormatError("Лист с ожидаемыми заголовками не найден")

def _meta_columns(header):
    idx = {}
    for i, cell in enumerate(header):
        if cell in META:
            idx[META[cell]] = i
    missing = set(META.values()) - set(idx)
    if missing:
        raise UnrecognizedFormatError(f"Не найдены колонки: {missing}")
    return idx

def _day_columns(date_row, header):
    days = []
    for i, cell in enumerate(header):
        if cell == LATE_HDR:
            d = _as_date(date_row[i]) if i < len(date_row) else None
            if d is None:
                continue
            arr = i + 1 if (i + 1 < len(header) and header[i + 1] == ARRIVAL_HDR) else None
            days.append((d, i, arr))
    return days

def read_records(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = _find_sheet(wb)
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 3:
        raise UnrecognizedFormatError("Недостаточно строк")
    date_row, header = rows[0], rows[1]
    meta = _meta_columns(header)
    days = _day_columns(date_row, header)
    records = []
    for row in rows[2:]:
        if row[meta["employee_no"]] in (None, ""):
            continue
        for d, lc, ac in days:
            late = row[lc] if lc < len(row) else None
            if late is None or str(late).strip() == "":
                continue
            try:
                late_min = float(late)
            except (TypeError, ValueError):
                continue
            records.append({
                "employee_no": row[meta["employee_no"]],
                "first_name": row[meta["first_name"]],
                "last_name": row[meta["last_name"]],
                "city": row[meta["city"]],
                "route": row[meta["route"]],
                "date": d.isoformat(),
                "weekday": _weekday(d),
                "late_min": late_min,
                "arrival": _fmt_time(row[ac]) if ac is not None and ac < len(row) else None,
            })
    return records
