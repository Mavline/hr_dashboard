import csv
import openpyxl
import pytest
from src import export

ROWS = [
    {"key": ["פתח תקווה"], "cases": 28, "total_late": 560,
     "avg_late": 20.0, "employees": 28, "routes": 1},
]

def test_write_xlsx(tmp_path):
    p = tmp_path / "out.xlsx"
    export.write(str(p), ROWS, "route", "xlsx")
    wb = openpyxl.load_workbook(p)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    assert headers == ["Route", "Cases", "Total Late (min)", "Avg Late (min)", "Employees", "Routes"]
    assert ws.max_row == 2  # шапка + 1 строка

def test_write_csv(tmp_path):
    p = tmp_path / "out.csv"
    export.write(str(p), ROWS, "route", "csv")
    with open(p, encoding="utf-8-sig", newline="") as f:
        data = list(csv.reader(f))
    assert data[0][0] == "Route"
    assert data[1][0] == "פתח תקווה"
    assert data[1][1] == "28"

def test_write_unknown_fmt(tmp_path):
    p = tmp_path / "out.txt"
    with pytest.raises(ValueError):
        export.write(str(p), ROWS, "route", "txt")

def test_write_unknown_view(tmp_path):
    p = tmp_path / "out.xlsx"
    with pytest.raises(ValueError):
        export.write(str(p), ROWS, "unknown", "xlsx")
